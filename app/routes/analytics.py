"""
Rutas de analytics y reportes
Gr√°ficas y estad√≠sticas avanzadas
"""
import datetime
import json
import io
import tempfile
import os
import locale
from flask import Blueprint, render_template, jsonify, make_response
from app.auth.auth_service import login_required, get_current_user, premium_required
from app.middleware.auth_middleware import require_auth, get_current_user_uid
from app.services.crop_service import CropService

def format_spanish_number(value, decimals=2):
    """Formatear n√∫mero con separador decimal espa√±ol (coma)"""
    try:
        if value is None:
            return "0"
        formatted = f"{float(value):.{decimals}f}"
        return formatted.replace('.', ',')
    except (ValueError, TypeError):
        return str(value)

analytics_bp = Blueprint('analytics', __name__)

# Guardia a nivel de blueprint: asegura autenticaci√≥n antes de cualquier handler
@analytics_bp.before_request
def _ensure_auth_analytics_bp():
    # Permitir preflight
    from flask import request as _req, redirect as _redirect, url_for as _url_for, session as _session, g as _g, current_app as _ca
    if _req.method == 'OPTIONS':
        return None
    # No interceptar endpoints API del propio blueprint; dejar que @require_auth maneje JSON/401
    try:
        if (_req.endpoint or '').endswith('analytics.api_chart_data') or _req.path.endswith('/api/chart-data'):
            return None
    except Exception:
        pass
    try:
        print(f"[analytics.before_request] method={_req.method} path={_req.path} args={dict(_req.args)} cookies={list(_req.cookies.keys())}")
    except Exception:
        pass
    # Resolver UID si ya fue establecido por middleware; si no, comprobar se√±ales de auth
    uid = None
    try:
        uid = (_g.current_user.get('uid') if getattr(_g, 'current_user', None) else None)
    except Exception:
        uid = None
    if not uid:
        uid = _session.get('user_uid')

    # Reconstrucci√≥n temprana desde cookies de respaldo si falta sesi√≥n pero hay cookies
    if not uid:
        try:
            backup_uid = _req.cookies.get('huerto_user_uid')
            backup_data = _req.cookies.get('huerto_user_data')
            if backup_uid:
                _session.permanent = True
                _session['user_uid'] = backup_uid
                _session['is_authenticated'] = True
                # Cargar m√≠nimos desde cookie JSON si existe
                try:
                    import json as _json
                    bd = _json.loads(backup_data) if backup_data else {}
                except Exception:
                    bd = {}
                _session['user'] = {
                    'uid': backup_uid,
                    'email': bd.get('email'),
                    'name': bd.get('name') or (bd.get('email') or 'Usuario'),
                    'plan': bd.get('plan', 'gratuito')
                }
                _session.modified = True
                # Exponer en g para capas superiores
                try:
                    _g.current_user = _session['user']
                except Exception:
                    pass
                uid = backup_uid
                try:
                    print("[analytics.before_request] Sesi√≥n reconstruida desde cookies de respaldo")
                except Exception:
                    pass
        except Exception:
            pass
    # Permitir token de desarrollo en local
    dev_token = _req.args.get('dev_token')
    if dev_token == 'dev_123_local' and (_ca.config.get('ENV') == 'development' or _ca.debug):
        try:
            print("[analytics.before_request] dev_token v√°lido en desarrollo; permitiendo acceso")
        except Exception:
            pass
        return None
    # Importante: no redirigir desde aqu√≠. Dejamos TODA la decisi√≥n al decorador @require_auth.
    # Este hook solo intenta reconstrucci√≥n temprana si hay cookies de respaldo y registra trazas.
    # Si no hay sesi√≥n/UID, continuamos y que el decorador gestione 401/redirects correctamente.
    return None

@analytics_bp.route('/')
def analytics_dashboard():
    """Dashboard de analytics (requiere autenticaci√≥n)."""
    from flask import current_app
    # Trazas de diagn√≥stico: cookies, sesi√≥n y UID detectado
    try:
        from flask import request as _req, session as _session, g as _g
        print(f"[/analytics/] IN - cookies={list((_req.cookies or {}).keys())} has_session={bool(_session.get('is_authenticated'))} session_uid={_session.get('user_uid')} g_uid={(getattr(_g, 'current_user', {}) or {}).get('uid')}")
    except Exception as _e:
        try:
            print(f"[/analytics/] IN - error trazas: {_e}")
        except Exception:
            pass
    
    # Obtener usuario autenticado con fallback robusto
    user = get_current_user()
    from flask import session as _session, request as _req
    user_uid = (
        get_current_user_uid()
        or _session.get('user_uid')
        or _req.cookies.get('huerto_user_uid')
        or _req.args.get('uid')
        or _req.args.get('user_uid')
    )
    try:
        print(f"[/analytics/] UID resuelto: {user_uid}")
    except Exception:
        pass
    # Evitar redirecciones falsas: si no hay UID pero hay se√±ales claras de auth, continuar
    if not user_uid:
        has_auth_header = bool(_req.headers.get('Authorization'))
        has_firebase_cookie = 'firebase_id_token' in _req.cookies
        has_user_cookie = 'huerto_user_uid' in _req.cookies or 'huerto_user_data' in _req.cookies
        came_from_auth = _req.args.get('from') in ("login", "register")
        try:
            print(f"[/analytics/] Sin UID. se√±ales: auth_header={has_auth_header} firebase_cookie={has_firebase_cookie} user_cookie={has_user_cookie} from_auth={came_from_auth}")
        except Exception:
            pass
        if not (has_auth_header or has_firebase_cookie or has_user_cookie or came_from_auth):
            from flask import redirect, url_for
            return redirect(url_for('main.onboarding'))
    
    from flask import current_app
    crop_service = CropService(current_app.db)
    
    # Obtener datos del usuario
    cultivos = crop_service.get_user_crops(user_uid)
    total_kilos, total_beneficios = crop_service.get_user_totals(user_uid)
    
    # Preparar datos detallados para cada cultivo
    cultivos_detallados = []
    for cultivo in cultivos:
        # Calcular totales de producci√≥n
        total_kilos_cultivo = sum(p.get('kilos', 0) for p in cultivo.get('produccion_diaria', []))
        total_unidades_cultivo = sum(p.get('unidades', 0) for p in cultivo.get('produccion_diaria', []))
        
        # Calcular peso promedio por unidad
        peso_por_unidad = (total_kilos_cultivo / total_unidades_cultivo) if total_unidades_cultivo > 0 else 0
        
        # Calcular beneficio
        precio_kilo = cultivo.get('precio_por_kilo', 0)
        beneficio = total_kilos_cultivo * precio_kilo
        
        # Calcular d√≠as de cultivo
        dias_cultivo = None
        if 'fecha_siembra' in cultivo and cultivo['fecha_siembra']:
            fecha_inicio = cultivo['fecha_siembra']
            fecha_fin = cultivo.get('fecha_cosecha', datetime.datetime.now())
            if hasattr(fecha_inicio, 'date'):
                fecha_inicio = fecha_inicio.date()
            if hasattr(fecha_fin, 'date'):
                fecha_fin = fecha_fin.date()
            elif isinstance(fecha_fin, datetime.datetime):
                fecha_fin = fecha_fin.date()
            
            try:
                dias_cultivo = (fecha_fin - fecha_inicio).days
            except:
                dias_cultivo = None
        
        # Contar abonos
        total_abonos = len(cultivo.get('abonos', []))
        
        cultivos_detallados.append({
            'nombre': cultivo['nombre'],
            'color_cultivo': cultivo.get('color_cultivo', '#28a745'),  # Incluir color del cultivo
            'activo': cultivo.get('activo', True),
            'numero_plantas': cultivo.get('numero_plantas', 1),
            'dias_cultivo': dias_cultivo,
            'total_unidades': total_unidades_cultivo,
            'total_kilos': total_kilos_cultivo,
            'peso_por_unidad': peso_por_unidad,
            'precio_por_kilo': precio_kilo,
            'beneficio': beneficio,
            'total_abonos': total_abonos
        })
    
    # Preparar datos para gr√°ficas (compatibilidad)
    datos_cultivos = {}
    for cultivo_det in cultivos_detallados:
        datos_cultivos[cultivo_det['nombre']] = {
            'kilos': cultivo_det['total_kilos'],
            'beneficio': cultivo_det['beneficio']
        }
    
    return render_template('analytics/dashboard.html',
                         datos_cultivos=datos_cultivos,
                         cultivos_detallados=cultivos_detallados,
                         total_kilos=total_kilos,
                         total_beneficios=total_beneficios,
                         user_uid=user_uid)

@analytics_bp.route('/advanced')
@require_auth
def advanced():
    """Analytics avanzados con autenticaci√≥n segura"""
    from flask import current_app
    
    # Obtener UID del usuario autenticado de forma segura
    user_uid = get_current_user_uid()
    user = get_current_user()
    
    from flask import current_app
    crop_service = CropService(current_app.db)
    
    # Obtener datos del usuario autenticado
    cultivos = crop_service.get_user_crops(user_uid)
    total_kilos, total_beneficios = crop_service.get_user_totals(user_uid)
    
    # Verificar plan del usuario
    plan = user.get('plan', 'gratuito')
    
    # Preparar datos avanzados para analytics
    # 1. An√°lisis temporal por meses
    monthly_data = {}
    for cultivo in cultivos:
        for produccion in cultivo.get('produccion_diaria', []):
            fecha = produccion['fecha']
            month_key = f"{fecha.year}-{fecha.month:02d}"
            if month_key not in monthly_data:
                monthly_data[month_key] = {'kilos': 0, 'beneficio': 0}
            monthly_data[month_key]['kilos'] += produccion['kilos']
            monthly_data[month_key]['beneficio'] += produccion['kilos'] * cultivo['precio_por_kilo']
    
    # 2. Ranking de cultivos m√°s rentables
    cultivos_ranking = sorted(cultivos, key=lambda x: x.get('beneficio_total', 0), reverse=True)
    
    # 3. Proyecciones (simuladas para demo)
    proyeccion_anual = total_beneficios * 2.5  # Proyecci√≥n optimista
    
    return render_template('analytics/advanced.html', 
                         cultivos=cultivos,
                         monthly_data=monthly_data,
                         cultivos_ranking=cultivos_ranking[:5],  # Top 5
                         total_kilos=total_kilos,
                         total_beneficios=total_beneficios,
                         proyeccion_anual=proyeccion_anual,
                         demo_mode=False,
                         plan=plan)

@analytics_bp.route('/api/chart-data')
def api_chart_data():
    """API para datos de gr√°ficas (requiere autenticaci√≥n)."""
    from flask import current_app, request, session as _session
    try:
        print(f"[/analytics/api/chart-data] IN - cookies={list((request.cookies or {}).keys())} session_uid={_session.get('user_uid')} g_uid={get_current_user_uid()}")
    except Exception:
        pass
    
    # Resolver UID de forma robusta: contexto -> sesi√≥n -> cookie -> par√°metro
    user_uid = (
        get_current_user_uid()
        or _session.get('user_uid')
        or request.cookies.get('huerto_user_uid')
        or request.args.get('uid')
        or request.args.get('user_uid')
    )
    # Fallback: intentar extraer UID del Referer (cuando venimos de /analytics/?uid=...)
    if not user_uid:
        try:
            ref = request.headers.get('Referer')
            if ref:
                from urllib.parse import urlparse, parse_qs
                q = parse_qs(urlparse(ref).query)
                user_uid = (q.get('uid') or q.get('user_uid') or [None])[0]
                if user_uid:
                    print(f"[/analytics/api/chart-data] UID desde Referer: {user_uid}")
        except Exception:
            pass
    try:
        print(f"[/analytics/api/chart-data] UID resuelto: {user_uid}")
    except Exception:
        pass
    if not user_uid:
        # Evitar 302 en APIs; responder JSON 401 para que el frontend pueda reaccionar
        return jsonify({'error': 'no-auth', 'redirect': '/onboarding'}), 401
    crop_service = CropService(current_app.db)
    cultivos = crop_service.get_user_crops(user_uid)
    
    # Preparar datos para Chart.js
    labels = []
    kilos_data = []
    beneficios_data = []
    colors = []  # Array de colores de cultivos
    
    for cultivo in cultivos:
        labels.append(cultivo['nombre'])
        
        kilos_total = sum(p.get('kilos', 0) for p in cultivo.get('produccion_diaria', []))
        precio_kilo = cultivo.get('precio_por_kilo', 0)
        
        kilos_data.append(kilos_total)
        beneficios_data.append(kilos_total * precio_kilo)
        colors.append(cultivo.get('color_cultivo', '#28a745'))  # Color espec√≠fico del cultivo
    
    return jsonify({
        'labels': labels,
        'colors': colors,  # Incluir colores espec√≠ficos de cada cultivo
        'datasets': [
            {
                'label': 'Kilogramos producidos',
                'data': kilos_data,
                'backgroundColor': colors,  # Usar colores espec√≠ficos
                'borderColor': colors,
                'borderWidth': 2
            },
            {
                'label': 'Beneficios (‚Ç¨)',
                'data': beneficios_data,
                'backgroundColor': colors,  # Usar colores espec√≠ficos
                'borderColor': colors,
                'borderWidth': 2
            }
        ]
    })

@analytics_bp.route('/export/csv')
@require_auth
def export_csv():
    """Exportar datos detallados a CSV (requiere autenticaci√≥n)."""
    from flask import current_app
    import csv
    import io
    
    user_uid = get_current_user_uid()
    crop_service = CropService(current_app.db)
    cultivos = crop_service.get_user_crops(user_uid)
    filename = f"huerto_detallado_{user_uid[:8]}.csv"
    
    # Crear CSV en memoria
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Cabeceras con todos los datos disponibles
    writer.writerow([
        'Cultivo', 'Estado', 'Fecha Siembra', 'Fecha Cosecha', 'D√≠as Cultivo',
        'N√∫mero Plantas', 'Precio/kg (‚Ç¨)', 'Total Unidades', 'Total Kilos',
        'Peso/Unidad (kg)', 'Beneficio Total (‚Ç¨)', 'Total Abonos', 'Rentabilidad (%)'
    ])
    
    # Calcular totales para rentabilidad
    total_beneficios = sum(
        sum(p.get('kilos', 0) for p in cultivo.get('produccion_diaria', [])) * cultivo.get('precio_por_kilo', 0)
        for cultivo in cultivos
    )
    
    # Datos detallados
    for cultivo in cultivos:
        # Calcular m√©tricas
        total_kilos = sum(p.get('kilos', 0) for p in cultivo.get('produccion_diaria', []))
        total_unidades = sum(p.get('unidades', 0) for p in cultivo.get('produccion_diaria', []))
        peso_por_unidad = (total_kilos / total_unidades) if total_unidades > 0 else 0
        beneficio = total_kilos * cultivo.get('precio_por_kilo', 0)
        rentabilidad = (beneficio / total_beneficios * 100) if total_beneficios > 0 else 0
        
        # Calcular d√≠as de cultivo
        dias_cultivo = ""
        if cultivo.get('fecha_siembra') and cultivo.get('fecha_cosecha'):
            delta = cultivo['fecha_cosecha'] - cultivo['fecha_siembra']
            dias_cultivo = delta.days
        elif cultivo.get('fecha_siembra'):
            delta = datetime.datetime.now() - cultivo['fecha_siembra']
            dias_cultivo = delta.days
        
        writer.writerow([
            cultivo['nombre'],
            'Activo' if cultivo.get('activo', True) else 'Finalizado',
            cultivo['fecha_siembra'].strftime('%Y-%m-%d') if cultivo.get('fecha_siembra') else '',
            cultivo['fecha_cosecha'].strftime('%Y-%m-%d') if cultivo.get('fecha_cosecha') else '',
            dias_cultivo,
            cultivo.get('numero_plantas', 1),
            format_spanish_number(cultivo.get('precio_por_kilo', 0), 2),
            total_unidades,
            format_spanish_number(total_kilos, 1),
            format_spanish_number(peso_por_unidad, 3) if peso_por_unidad > 0 else "0",
            format_spanish_number(beneficio, 2),
            len(cultivo.get('abonos', [])),
            format_spanish_number(rentabilidad, 1)
        ])
    
    # Crear respuesta HTTP
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@analytics_bp.route('/export/json')
@require_auth
def export_json():
    """Exportar datos completos a JSON (requiere autenticaci√≥n)."""
    from flask import current_app
    
    user_uid = get_current_user_uid()
    crop_service = CropService(current_app.db)
    cultivos = crop_service.get_user_crops(user_uid)
    total_kilos, total_beneficios = crop_service.get_user_totals(user_uid)
    
    # Preparar datos completos para exportar
    export_data = {
        'resumen': {
            'total_cultivos': len(cultivos),
            'total_kilos': total_kilos,
            'total_beneficios': total_beneficios,
            'fecha_exportacion': datetime.datetime.now().isoformat(),
            'usuario': user_uid
        },
        'cultivos_detallados': []
    }
    
    for cultivo in cultivos:
        # Calcular m√©tricas detalladas
        total_kilos_cultivo = sum(p.get('kilos', 0) for p in cultivo.get('produccion_diaria', []))
        total_unidades_cultivo = sum(p.get('unidades', 0) for p in cultivo.get('produccion_diaria', []))
        peso_por_unidad = (total_kilos_cultivo / total_unidades_cultivo) if total_unidades_cultivo > 0 else 0
        beneficio = total_kilos_cultivo * cultivo.get('precio_por_kilo', 0)
        
        # Calcular d√≠as de cultivo
        dias_cultivo = None
        if cultivo.get('fecha_siembra'):
            fecha_fin = cultivo.get('fecha_cosecha', datetime.datetime.now())
            try:
                dias_cultivo = (fecha_fin - cultivo['fecha_siembra']).days
            except:
                dias_cultivo = None
        
        cultivo_data = {
            'informacion_basica': {
                'nombre': cultivo['nombre'],
                'activo': cultivo.get('activo', True),
                'numero_plantas': cultivo.get('numero_plantas', 1),
                'precio_por_kilo': cultivo.get('precio_por_kilo', 0)
            },
            'fechas': {
                'fecha_siembra': cultivo['fecha_siembra'].isoformat() if cultivo.get('fecha_siembra') else None,
                'fecha_cosecha': cultivo['fecha_cosecha'].isoformat() if cultivo.get('fecha_cosecha') else None,
                'dias_cultivo': dias_cultivo
            },
            'produccion': {
                'total_kilos': total_kilos_cultivo,
                'total_unidades': total_unidades_cultivo,
                'peso_por_unidad': peso_por_unidad,
                'beneficio_total': beneficio,
                'produccion_diaria': [
                    {
                        'fecha': p['fecha'].isoformat(),
                        'kilos': p.get('kilos', 0),
                        'unidades': p.get('unidades', 0),
                        'peso_unitario': p.get('kilos', 0) / p.get('unidades', 1) if p.get('unidades', 0) > 0 else 0
                    } for p in cultivo.get('produccion_diaria', [])
                ]
            },
            'mantenimiento': {
                'total_abonos': len(cultivo.get('abonos', [])),
                'abonos_detallados': [
                    {
                        'fecha': a['fecha'].isoformat() if 'fecha' in a else None,
                        'descripcion': a.get('descripcion', '')
                    } for a in cultivo.get('abonos', [])
                ]
            },
            'estadisticas': {
                'rentabilidad_porcentaje': (beneficio / total_beneficios * 100) if total_beneficios > 0 else 0,
                'productividad_por_planta': total_kilos_cultivo / cultivo.get('numero_plantas', 1),
                'beneficio_por_dia': beneficio / dias_cultivo if dias_cultivo and dias_cultivo > 0 else 0
            }
        }
        export_data['cultivos_detallados'].append(cultivo_data)
    
    filename = f"huerto_completo_{user_uid[:8]}.json"
    
    # Crear respuesta HTTP
    response = make_response(json.dumps(export_data, indent=2, ensure_ascii=False))
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@analytics_bp.route('/export/excel')
@require_auth
def export_excel():
    """Exportar datos a Excel con m√∫ltiples hojas (requiere autenticaci√≥n)."""
    from flask import current_app
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    
    user_uid = get_current_user_uid()
    crop_service = CropService(current_app.db)
    cultivos = crop_service.get_user_crops(user_uid)
    total_kilos, total_beneficios = crop_service.get_user_totals(user_uid)
    filename = f"huerto_analisis_{user_uid[:8]}.xlsx"
    
    # Crear libro Excel
    wb = Workbook()
    
    # === HOJA 1: RESUMEN ===
    ws1 = wb.active
    ws1.title = "Resumen General"
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    
    # T√≠tulo principal
    ws1['A1'] = "üå± HuertoRentable - An√°lisis Completo"
    ws1['A1'].font = Font(bold=True, size=16, color="198754")
    ws1.merge_cells('A1:D1')
    
    # Resumen general
    ws1['A3'] = "Resumen General"
    ws1['A3'].font = header_font
    ws1['A3'].fill = header_fill
    ws1.merge_cells('A3:B3')
    
    ws1['A4'] = "Total Cultivos:"
    ws1['B4'] = len(cultivos)
    ws1['A5'] = "Producci√≥n Total:"
    ws1['B5'] = f"{total_kilos:.1f} kg"
    ws1['A6'] = "Beneficios Totales:"
    ws1['B6'] = f"{format_spanish_number(total_beneficios, 2)} ‚Ç¨"
    ws1['A7'] = "Fecha Reporte:"
    ws1['B7'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # === HOJA 2: DETALLE CULTIVOS COMPLETO ===
    ws2 = wb.create_sheet(title="Detalle Cultivos Completo")
    
    # Encabezados expandidos con todos los datos
    headers = ['Cultivo', 'Estado', 'Fecha Siembra', 'Fecha Cosecha', 'D√≠as Cultivo',
               'N√∫mero Plantas', 'Precio/kg (‚Ç¨)', 'Total Unidades', 'Total Kilos', 
               'Peso/Unidad (kg)', 'Beneficio Total (‚Ç¨)', 'Total Abonos', 'Rentabilidad (%)']
    
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Datos detallados de cultivos
    for row, cultivo in enumerate(cultivos, 2):
        # Calcular m√©tricas
        total_kilos_cultivo = sum(p.get('kilos', 0) for p in cultivo.get('produccion_diaria', []))
        total_unidades_cultivo = sum(p.get('unidades', 0) for p in cultivo.get('produccion_diaria', []))
        peso_por_unidad = (total_kilos_cultivo / total_unidades_cultivo) if total_unidades_cultivo > 0 else 0
        beneficio = total_kilos_cultivo * cultivo.get('precio_por_kilo', 0)
        rentabilidad = (beneficio / total_beneficios * 100) if total_beneficios > 0 else 0
        
        # Calcular d√≠as de cultivo
        dias_cultivo = ""
        if cultivo.get('fecha_siembra'):
            fecha_fin = cultivo.get('fecha_cosecha', datetime.datetime.now())
            try:
                dias_cultivo = (fecha_fin - cultivo['fecha_siembra']).days
            except:
                dias_cultivo = ""
        
        # Llenar datos
        ws2.cell(row=row, column=1, value=cultivo['nombre'])
        ws2.cell(row=row, column=2, value='Activo' if cultivo.get('activo', True) else 'Finalizado')
        ws2.cell(row=row, column=3, value=cultivo['fecha_siembra'].strftime('%Y-%m-%d') if cultivo.get('fecha_siembra') else '')
        ws2.cell(row=row, column=4, value=cultivo['fecha_cosecha'].strftime('%Y-%m-%d') if cultivo.get('fecha_cosecha') else '')
        ws2.cell(row=row, column=5, value=dias_cultivo)
        ws2.cell(row=row, column=6, value=cultivo.get('numero_plantas', 1))
        ws2.cell(row=row, column=7, value=cultivo.get('precio_por_kilo', 0))
        ws2.cell(row=row, column=8, value=total_unidades_cultivo)
        ws2.cell(row=row, column=9, value=total_kilos_cultivo)
        ws2.cell(row=row, column=10, value=peso_por_unidad)
        ws2.cell(row=row, column=11, value=beneficio)
        ws2.cell(row=row, column=12, value=len(cultivo.get('abonos', [])))
        ws2.cell(row=row, column=13, value=rentabilidad)
    
    # Ajustar anchos de columna
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws2.column_dimensions[column].width = adjusted_width
    
    # === HOJA 3: PRODUCCI√ìN DETALLADA ===
    if cultivos and any(cultivo.get('produccion_diaria', []) for cultivo in cultivos):
        ws3 = wb.create_sheet(title="Producci√≥n Diaria")
        
        ws3['A1'] = "Producci√≥n Detallada por D√≠a"
        ws3['A1'].font = header_font
        ws3['A1'].fill = header_fill
        ws3.merge_cells('A1:E1')
        
        headers_prod = ['Cultivo', 'Fecha', 'Kilos', 'Unidades', 'Peso/Unidad (kg)']
        for col, header in enumerate(headers_prod, 1):
            cell = ws3.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 3
        for cultivo in cultivos:
            for produccion in cultivo.get('produccion_diaria', []):
                kilos = produccion.get('kilos', 0)
                unidades = produccion.get('unidades', 0)
                peso_unitario = (kilos / unidades) if unidades > 0 else 0
                
                ws3.cell(row=row, column=1, value=cultivo['nombre'])
                ws3.cell(row=row, column=2, value=produccion['fecha'].strftime('%Y-%m-%d'))
                ws3.cell(row=row, column=3, value=kilos)
                ws3.cell(row=row, column=4, value=unidades)
                ws3.cell(row=row, column=5, value=peso_unitario)
                row += 1
                
        # Ajustar anchos de esta hoja tambi√©n
        for col in ws3.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws3.column_dimensions[column].width = adjusted_width
    
    # === HOJA 4: HISTORIAL DE ABONOS ===
    if cultivos and any(cultivo.get('abonos', []) for cultivo in cultivos):
        ws4 = wb.create_sheet(title="Historial Abonos")
        
        ws4['A1'] = "Historial de Abonos y Mantenimiento"
        ws4['A1'].font = header_font
        ws4['A1'].fill = header_fill
        ws4.merge_cells('A1:C1')
        
        headers_abonos = ['Cultivo', 'Fecha', 'Descripci√≥n']
        for col, header in enumerate(headers_abonos, 1):
            cell = ws4.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 3
        for cultivo in cultivos:
            for abono in cultivo.get('abonos', []):
                ws4.cell(row=row, column=1, value=cultivo['nombre'])
                ws4.cell(row=row, column=2, value=abono['fecha'].strftime('%Y-%m-%d') if 'fecha' in abono else '')
                ws4.cell(row=row, column=3, value=abono.get('descripcion', ''))
                row += 1
                
        # Ajustar anchos
        for col in ws4.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Un poco m√°s ancho para descripciones
            ws4.column_dimensions[column].width = adjusted_width
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta HTTP
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@analytics_bp.route('/export/pdf')
@require_auth
def export_pdf():
    """Exportar reporte completo a PDF con reportlab (requiere autenticaci√≥n)."""
    from flask import current_app
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    
    user_uid = get_current_user_uid()
    crop_service = CropService(current_app.db)
    cultivos = crop_service.get_user_crops(user_uid)
    total_kilos, total_beneficios = crop_service.get_user_totals(user_uid)
    filename = f"huerto_reporte_{user_uid[:8]}.pdf"
    titulo_usuario = f"Usuario: {user_uid}"
    
    # Crear PDF en memoria
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Preparar estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#198754'),
        alignment=1  # Center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor('#198754')
    )
    
    # Contenido del PDF
    story = []
    
    # T√≠tulo principal
    story.append(Paragraph("üå± HuertoRentable", title_style))
    story.append(Paragraph("Reporte Completo de An√°lisis", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    # Informaci√≥n del usuario y fecha
    story.append(Paragraph(f"<b>{titulo_usuario}</b>", styles['Normal']))
    story.append(Paragraph(f"Fecha: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Resumen ejecutivo
    story.append(Paragraph("üìä Resumen Ejecutivo", heading_style))
    
    # Tabla de resumen
    resumen_data = [
        ['M√©trica', 'Valor'],
        ['Total de Cultivos', str(len(cultivos))],
        ['Producci√≥n Total', f"{total_kilos:.1f} kg"],
        ['Beneficios Totales', f"{format_spanish_number(total_beneficios, 2)} ‚Ç¨"],
        ['Beneficio Promedio', f"{format_spanish_number((total_beneficios/len(cultivos) if cultivos else 0), 2)} ‚Ç¨ por cultivo"]
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(resumen_table)
    story.append(Spacer(1, 20))
    
    # Detalle de cultivos
    if cultivos:
        story.append(Paragraph("üå± Detalle de Cultivos", heading_style))
        
        # Tabla de cultivos
        cultivos_data = [['Cultivo', 'Fecha Siembra', 'Plantas', 'Precio/kg', 'Kilos', 'Beneficio', 'Estado']]
        
        for cultivo in cultivos:
            cultivos_data.append([
                cultivo['nombre'],
                cultivo['fecha_siembra'].strftime('%d/%m/%Y'),
                str(cultivo.get('plantas_sembradas', 0)),
                f"{format_spanish_number(cultivo.get('precio_por_kilo', 0), 2)} ‚Ç¨",
                f"{cultivo.get('kilos_totales', 0):.1f} kg",
                f"{format_spanish_number(cultivo.get('beneficio_total', 0), 2)} ‚Ç¨",
                'Activo' if cultivo.get('activo', True) else 'Cosechado'
            ])
        
        cultivos_table = Table(cultivos_data, colWidths=[0.8*inch, 0.8*inch, 0.6*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.7*inch])
        cultivos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(cultivos_table)
        story.append(Spacer(1, 20))
        
        # An√°lisis de rentabilidad
        story.append(Paragraph("üíπ An√°lisis de Rentabilidad", heading_style))
        
        if cultivos:
            cultivo_top = max(cultivos, key=lambda x: x.get('beneficio_total', 0))
            story.append(Paragraph(f"<b>üèÜ Cultivo M√°s Rentable:</b> {cultivo_top['nombre']}", styles['Normal']))
            story.append(Paragraph(f"Beneficio: {format_spanish_number(cultivo_top.get('beneficio_total', 0), 2)} ‚Ç¨", styles['Normal']))
            story.append(Paragraph(f"Producci√≥n: {cultivo_top.get('kilos_totales', 0):.1f} kg", styles['Normal']))
            story.append(Spacer(1, 12))
            
            activos = len([c for c in cultivos if c.get('activo', True)])
            cosechados = len(cultivos) - activos
            story.append(Paragraph(f"<b>üìä Estad√≠sticas:</b>", styles['Normal']))
            story.append(Paragraph(f"‚Ä¢ Cultivos activos: {activos}", styles['Normal']))
            story.append(Paragraph(f"‚Ä¢ Cultivos cosechados: {cosechados}", styles['Normal']))
            if total_kilos > 0:
                story.append(Paragraph(f"‚Ä¢ Rentabilidad: {format_spanish_number((total_beneficios/total_kilos), 2)} ‚Ç¨/kg", styles['Normal']))
    else:
        story.append(Paragraph("No hay cultivos registrados", styles['Normal']))
    
    # Pie de p√°gina
    story.append(Spacer(1, 30))
    story.append(Paragraph("üå± <b>HuertoRentable</b> - Gesti√≥n inteligente de huertos rentables", styles['Normal']))
    story.append(Paragraph(f"Reporte generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    
    # Crear respuesta HTTP
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response
    
    return response
